from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
import sqlite3
import os
import sys
import time
import threading
import signal
from functools import wraps
from datetime import datetime, date
from werkzeug.security import generate_password_hash, check_password_hash
from audit import log_auditoria

_DB_SETTINGS = {
    "mode": "AUTO",  # "AUTO", "REDE_G", "LOCAL"
    "net_dir": r"G:\Triagem AITs",
    "cache_path": None,
    "last_check": 0
}

def get_local_db_path():
    if getattr(sys, 'frozen', False):
        return os.path.join(os.path.dirname(sys.executable), "triagem_ait.db")
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "triagem_ait.db")

def get_net_db_path():
    return os.path.join(_DB_SETTINGS["net_dir"], "triagem_ait.db")

def get_base_path():
    now = time.time()
    mode = _DB_SETTINGS.get("mode", "AUTO")
    
    if mode == "LOCAL":
        return get_local_db_path()
        
    if mode == "REDE_G":
        return get_net_db_path()
        
    # AUTO mode: com cache em memória de 30s
    if _DB_SETTINGS["cache_path"] and (now - _DB_SETTINGS["last_check"] < 30):
        return _DB_SETTINGS["cache_path"]
        
    net_dir = _DB_SETTINGS["net_dir"]
    net_db = get_net_db_path()
    resolved_path = None
    try:
        if os.path.exists(net_dir) and os.path.exists(net_db):
            resolved_path = net_db
    except Exception:
        resolved_path = None

    if not resolved_path:
        resolved_path = get_local_db_path()
            
    _DB_SETTINGS["cache_path"] = resolved_path
    _DB_SETTINGS["last_check"] = now
    return resolved_path

def get_resources_path():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

app = Flask(
    __name__,
    template_folder=os.path.join(get_resources_path(), "templates"),
    static_folder=os.path.join(get_resources_path(), "static")
)
app.secret_key = "triagem_ait_v11_secure_key"

def get_db_connection():
    conn = sqlite3.connect(get_base_path(), timeout=30.0)
    conn.execute("PRAGMA journal_mode = WAL;")
    conn.execute("PRAGMA synchronous = NORMAL;")
    conn.execute("PRAGMA busy_timeout = 30000;")
    conn.row_factory = sqlite3.Row
    return conn

def init_db_performance_indexes():
    try:
        conn = get_db_connection()
        conn.execute("CREATE INDEX IF NOT EXISTS idx_ait_numero ON ait(numero_ait);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_ait_talao ON ait(talao_id);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_ait_agente ON ait(agente_gcm_id);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_ait_data ON ait(data_ait);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_ait_status ON ait(status);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_taloes_agente ON taloes(agente_gcm_id);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_usuarios_username ON usuarios(username);")
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[WARN] Erro ao inicializar indices: {e}")

try:
    from scripts.backup_manager import iniciar_agendador_backups
    iniciar_agendador_backups()
except Exception as e:
    print(f"[BACKUP AGENDADOR AVISO]: {e}")

def with_db_retry(max_retries=5, delay=0.4):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            last_err = None
            for attempt in range(max_retries):
                try:
                    return f(*args, **kwargs)
                except sqlite3.OperationalError as e:
                    err_msg = str(e).lower()
                    if "locked" in err_msg or "busy" in err_msg or "io error" in err_msg or "disk" in err_msg:
                        last_err = e
                        _DB_SETTINGS["last_check"] = 0
                        time.sleep(delay * (attempt + 1))
                    else:
                        raise e
                except Exception as e:
                    last_err = e
                    _DB_SETTINGS["last_check"] = 0
                    time.sleep(delay * (attempt + 1))
            raise last_err
        return wrapper
    return decorator

# --- Auth & RBAC ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session:
            flash("Faça login para continuar.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if "user" not in session:
                flash("Faça login para continuar.", "warning")
                return redirect(url_for("login"))
            user_role = session["user"].get("setor")
            if user_role not in roles and user_role != "admin":
                flash("Acesso não autorizado para o seu setor.", "danger")
                return redirect(url_for("index"))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.template_filter('format_date')
def format_date(value):
    if not value:
        return ""
    try:
        dt = datetime.strptime(str(value).split(' ')[0], "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return value

@app.template_filter('format_month_year')
def format_month_year(value):
    if not value:
        return ""
    try:
        parts = str(value).split('-')
        if len(parts) == 2:
            return f"{parts[1]}/{parts[0]}"
        return value
    except Exception:
        return value

USER_NAMES_CACHE = {}

@app.template_filter('user_fullname')
def user_fullname(username):
    if not username:
        return ""
    if username in USER_NAMES_CACHE:
        return USER_NAMES_CACHE[username]
    try:
        conn = get_db_connection()
        user = conn.execute("SELECT nome_completo FROM usuarios WHERE username = ?", (username,)).fetchone()
        conn.close()
        if user and user["nome_completo"]:
            USER_NAMES_CACHE[username] = user["nome_completo"]
            return f"{user['nome_completo']} ({username})"
    except Exception:
        pass
    return username

@app.context_processor
def inject_globals():
    return {
        'datetime_now': datetime.now().strftime("%Y-%m-%d %H:%M"),
        'current_user': session.get('user')
    }

@app.before_request
def update_last_seen():
    if "user" in session:
        now_ts = time.time()
        last_update = session.get("_last_seen_ts", 0)
        if now_ts - last_update > 60:
            session["_last_seen_ts"] = now_ts
            try:
                conn = get_db_connection()
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                conn.execute("UPDATE usuarios SET ultimo_acesso = ? WHERE id = ?", (now_str, session["user"]["id"]))
                conn.commit()
                conn.close()
            except Exception:
                pass

@app.route("/quem_esta_logado")
@login_required
@role_required("admin", "dct")
def quem_esta_logado():
    conn = get_db_connection()
    users_active = conn.execute("""
        SELECT id, username, nome_completo, setor, vinculo, ultimo_acesso, ativo
        FROM usuarios
        WHERE ultimo_acesso IS NOT NULL
        ORDER BY ultimo_acesso DESC
    """).fetchall()
    conn.close()
    return render_template("quem_esta_logado.html", users_active=users_active)


STATUS_OPTIONS = ["DCT PROCESSAR", "PROCESSADO DCT", "CANCELADO", "AIT SUBSTITUIDA", "RENAINF"]

# --- API Status e Gerenciamento de Rede ---
@app.route("/api/status_rede")
def api_status_rede():
    net_dir = _DB_SETTINGS["net_dir"]
    net_db = get_net_db_path()
    current_mode = _DB_SETTINGS.get("mode", "AUTO")
    active_path = get_base_path()
    local_path = get_local_db_path()
    
    net_available = False
    latency_ms = None
    net_error = None
    
    start_t = time.time()
    try:
        if os.path.exists(net_dir) and os.path.exists(net_db):
            conn = sqlite3.connect(net_db, timeout=2.0)
            conn.execute("SELECT 1")
            conn.close()
            latency_ms = round((time.time() - start_t) * 1000, 1)
            net_available = True
    except Exception as e:
        net_error = str(e)

    is_using_net = (active_path == net_db)
    
    return jsonify({
        "status": "online" if (active_path and os.path.exists(active_path)) else "offline",
        "modo": current_mode,
        "usando_rede": is_using_net,
        "caminho_ativo": active_path,
        "rede_g_disponivel": net_available,
        "rede_g_caminho": net_db,
        "rede_g_latencia_ms": latency_ms,
        "rede_g_erro": net_error,
        "local_caminho": local_path
    }), 200

@app.route("/api/rede/conectar", methods=["POST"])
def api_rede_conectar():
    data = request.get_json(silent=True) or request.form or {}
    novo_modo = (data.get("modo") or "AUTO").upper().strip()
    if novo_modo not in ["AUTO", "REDE_G", "LOCAL"]:
        return jsonify({"sucesso": False, "mensagem": "Modo inválido. Escolha AUTO, REDE_G ou LOCAL."}), 400
        
    _DB_SETTINGS["mode"] = novo_modo
    _DB_SETTINGS["cache_path"] = None
    _DB_SETTINGS["last_check"] = 0
    
    active_path = get_base_path()
    net_db = get_net_db_path()
    
    if novo_modo == "REDE_G":
        if not (os.path.exists(_DB_SETTINGS["net_dir"]) and os.path.exists(net_db)):
            return jsonify({
                "sucesso": False,
                "mensagem": f"Pasta ou arquivo de banco na Rede G: ({net_db}) não encontrado. Verifique se o drive G:\\ está mapeado.",
                "modo": novo_modo,
                "caminho_ativo": active_path
            }), 404

    return jsonify({
        "sucesso": True,
        "mensagem": f"Conexão alterada para modo {novo_modo} com sucesso!",
        "modo": novo_modo,
        "caminho_ativo": active_path
    }), 200

@app.route("/api/rede/testar", methods=["POST", "GET"])
def api_rede_testar():
    net_dir = _DB_SETTINGS["net_dir"]
    net_db = get_net_db_path()
    start_t = time.time()
    
    if not os.path.exists(net_dir):
        return jsonify({
            "sucesso": False,
            "mensagem": f"Pasta '{net_dir}' não encontrada. Verifique se o drive de rede G:\\ está conectado."
        }), 404
        
    if not os.path.exists(net_db):
        return jsonify({
            "sucesso": False,
            "mensagem": f"Pasta '{net_dir}' localizada, porém o arquivo de banco '{net_db}' não existe."
        }), 404
        
    try:
        conn = sqlite3.connect(net_db, timeout=3.0)
        conn.execute("SELECT 1")
        conn.close()
        elapsed_ms = round((time.time() - start_t) * 1000, 1)
        return jsonify({
            "sucesso": True,
            "mensagem": f"Conexão com Rede G:\\Triagem AITs bem-sucedida! Latência: {elapsed_ms}ms",
            "latencia_ms": elapsed_ms,
            "caminho": net_db
        }), 200
    except Exception as e:
        return jsonify({
            "sucesso": False,
            "mensagem": f"Erro de acesso ao banco na Rede G:: {e}"
        }), 500

# --- Authentication Routes ---
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        conn = get_db_connection()
        user = conn.execute("SELECT * FROM usuarios WHERE username = ? AND ativo = 1", (username,)).fetchone()
        conn.close()

        if user and check_password_hash(user["senha_hash"], password):
            session["user"] = {
                "id": user["id"],
                "username": user["username"],
                "nome_completo": user["nome_completo"],
                "setor": user["setor"],
                "vinculo": user["vinculo"]
            }
            session["_last_seen_ts"] = time.time()
            log_auditoria(user["username"], user["setor"], "LOGIN_SUCESSO")
            flash(f"Bem-vindo(a), {user['nome_completo']}!", "success")
            return redirect(url_for("index"))
        else:
            log_auditoria(username, "DESCONHECIDO", "LOGIN_FALHA", justificativa="Senha ou usuário incorreto")
    conn = get_db_connection()
    triagem_users = conn.execute("SELECT username, nome_completo FROM usuarios WHERE setor IN ('transporte', 'triagem', 'setor_publico') AND username NOT IN ('admin', 'transporte') AND ativo = 1 ORDER BY nome_completo ASC").fetchall()
    dct_users = conn.execute("SELECT username, nome_completo FROM usuarios WHERE setor = 'dct' AND ativo = 1 ORDER BY nome_completo ASC").fetchall()
    admin_users = conn.execute("SELECT username, nome_completo FROM usuarios WHERE setor = 'admin' AND ativo = 1 ORDER BY nome_completo ASC").fetchall()
    conn.close()

    return render_template("login.html", triagem_users=triagem_users, dct_users=dct_users, admin_users=admin_users)

@app.route("/logout")
def logout():
    if "user" in session:
        log_auditoria(session["user"]["username"], session["user"]["setor"], "LOGOUT")
    session.pop("user", None)
    flash("Sessão encerrada com sucesso. O servidor Triagem AIT continua ativo na barra de tarefas (próximo ao relógio).", "info")
    return redirect(url_for("login"))

@app.route("/system/shutdown", methods=["POST", "GET"])
def system_shutdown():
    user = session.get("user", {})
    user_name = user.get("username", "DESCONHECIDO")
    user_setor = user.get("setor", "SISTEMA")
    log_auditoria(user_name, user_setor, "SHUTDOWN_SERVIDOR", justificativa="Encerramento manual do servidor via interface")
    
    # Gera backup seguro pré-encerramento
    try:
        from scripts.backup_manager import criar_backup
        criar_backup(tipo="SHUTDOWN")
    except Exception:
        pass

    def kill_process():
        # Fecha a guia ativa do navegador se a janela em foco for o navegador
        try:
            import ctypes
            user32 = ctypes.windll.user32
            hwnd = user32.GetForegroundWindow()
            if hwnd:
                length = user32.GetWindowTextLengthW(hwnd)
                title = ""
                if length > 0:
                    buf = ctypes.create_unicode_buffer(length + 1)
                    user32.GetWindowTextW(hwnd, buf, length + 1)
                    title = buf.value.lower()
                
                # Proteção: NUNCA envia comando se a janela em foco for a IDE, editor ou terminal
                if not any(k in title for k in ["antigravity", "visual studio", "vscode", "terminal", "powershell", "cmd", "python"]):
                    VK_CONTROL = 0x11
                    VK_W = 0x57
                    KEYEVENTF_KEYUP = 0x0002
                    user32.keybd_event(VK_CONTROL, 0, 0, 0)
                    user32.keybd_event(VK_W, 0, 0, 0)
                    user32.keybd_event(VK_W, 0, KEYEVENTF_KEYUP, 0)
                    user32.keybd_event(VK_CONTROL, 0, KEYEVENTF_KEYUP, 0)
        except Exception:
            pass

        time.sleep(0.3)
        # Finaliza o processo do servidor
        try:
            os._exit(0)
        except Exception:
            pass
        try:
            sys.exit(0)
        except Exception:
            pass
        try:
            os.kill(os.getpid(), signal.SIGTERM)
        except Exception:
            pass
        
    threading.Thread(target=kill_process, daemon=True).start()
    return jsonify({"status": "shutdown", "message": "Servidor encerrado com sucesso!"}), 200

# --- Dashboard Principal ---
@app.route("/")
@login_required
def index():
    conn = get_db_connection()
    stats = {}
    kpis = {
        'digitados_hoje': 0,
        'pendentes_dct': 0,
        'aits_faltantes': 0,
        'ultimo_backup_str': 'Nenhum backup',
        'ultimo_backup_size': ''
    }
    recent_records = []
    
    try:
        stats['total'] = conn.execute("SELECT COUNT(*) FROM ait").fetchone()[0]
        stats['dct_processar'] = conn.execute("SELECT COUNT(*) FROM ait WHERE status = 'DCT PROCESSAR'").fetchone()[0]
        stats['cancelado'] = conn.execute("SELECT COUNT(*) FROM ait WHERE status = 'CANCELADO'").fetchone()[0]
        stats['substituida'] = conn.execute("SELECT COUNT(*) FROM ait WHERE status = 'AIT SUBSTITUIDA'").fetchone()[0]
        stats['renainf'] = conn.execute("SELECT COUNT(*) FROM ait WHERE status = 'RENAINF'").fetchone()[0]
        
        # KPIs
        today_str = datetime.today().strftime("%Y-%m-%d")
        kpis['digitados_hoje'] = conn.execute("SELECT COUNT(*) FROM ait WHERE data_ait = ? OR date(criado_em) = ?", (today_str, today_str)).fetchone()[0]
        kpis['pendentes_dct'] = stats['dct_processar']
        
        # Calculate missing AIT count
        total_taloes_aits = conn.execute("SELECT SUM(quantidade_calculada) FROM taloes").fetchone()[0] or 0
        kpis['aits_faltantes'] = max(0, total_taloes_aits - stats['total'])
        
        recent_records = conn.execute("SELECT * FROM ait ORDER BY id DESC LIMIT 5").fetchall()
    except Exception:
        stats = {'total': 0, 'dct_processar': 0, 'cancelado': 0, 'substituida': 0, 'renainf': 0}
    finally:
        conn.close()

    try:
        backups_list = listar_backups()
        if backups_list:
            last_b = backups_list[0]
            kpis['ultimo_backup_str'] = last_b['data_hora']
            kpis['ultimo_backup_size'] = f"{last_b['size_mb']} MB"
    except Exception:
        pass

    return render_template("index.html", stats=stats, kpis=kpis, recent_records=recent_records)

# --- Agentes e GCMs ---
@app.route("/agentes")
@login_required
def agentes():
    conn = get_db_connection()
    list_agentes = conn.execute("SELECT * FROM agentes_gcm ORDER BY nome_completo ASC").fetchall()
    conn.close()
    return render_template("agentes.html", agentes=list_agentes)

@app.route("/agentes/criar", methods=["POST"])
@login_required
@role_required("dct", "admin")
@with_db_retry()
def agentes_criar():
    nome = request.form.get("nome_completo", "").strip()
    matricula = request.form.get("matricula", "").strip()
    categoria = request.form.get("categoria", "").strip()
    unidade = request.form.get("unidade_setor", "").strip()
    user_name = session["user"]["username"]

    if not nome or not matricula or not categoria:
        flash("Nome, Matrícula e Categoria são obrigatórios.", "danger")
        return redirect(url_for("agentes"))

    conn = get_db_connection()
    try:
        conn.execute("""
        INSERT INTO agentes_gcm (nome_completo, matricula, categoria, unidade_setor, criado_por)
        VALUES (?, ?, ?, ?, ?)
        """, (nome, matricula, categoria, unidade, user_name))
        conn.commit()
        log_auditoria(user_name, session["user"]["setor"], "CRIAR_AGENTE", "agentes_gcm", depois=f"{nome} ({matricula})")
        flash(f"Servidor '{nome}' cadastrado com sucesso!", "success")
    except sqlite3.IntegrityError:
        flash(f"Erro: A matrícula '{matricula}' já existe.", "danger")
    finally:
        conn.close()

    return redirect(url_for("agentes"))

@app.route("/agentes/editar/<int:agente_id>", methods=["POST"])
@login_required
@role_required("dct", "admin")
@with_db_retry()
def agentes_editar(agente_id):
    nome = request.form.get("nome_completo", "").strip()
    matricula = request.form.get("matricula", "").strip()
    categoria = request.form.get("categoria", "").strip()
    unidade = request.form.get("unidade_setor", "").strip()
    situacao = request.form.get("situacao", "ATIVO").strip()
    user_name = session["user"]["username"]

    if not nome or not matricula or not categoria:
        flash("Nome, Matrícula e Categoria são obrigatórios.", "danger")
        return redirect(url_for("agentes"))

    conn = get_db_connection()
    try:
        dup = conn.execute("SELECT id FROM agentes_gcm WHERE matricula = ? AND id != ?", (matricula, agente_id)).fetchone()
        if dup:
            flash(f"Erro: A matrícula '{matricula}' já pertence a outro servidor.", "danger")
            conn.close()
            return redirect(url_for("agentes"))

        old_agente = conn.execute("SELECT * FROM agentes_gcm WHERE id = ?", (agente_id,)).fetchone()
        antes_txt = f"{old_agente['nome_completo']} (Mat: {old_agente['matricula']}) | {old_agente['categoria']} | {old_agente['situacao']}" if old_agente else None
        depois_txt = f"{nome} (Mat: {matricula}) | {categoria} | {situacao}"

        conn.execute("""
            UPDATE agentes_gcm 
            SET nome_completo = ?, matricula = ?, categoria = ?, unidade_setor = ?, situacao = ?
            WHERE id = ?
        """, (nome, matricula, categoria, unidade, situacao, agente_id))
        conn.commit()
        log_auditoria(user_name, session["user"]["setor"], "EDITAR_AGENTE", "agentes_gcm", agente_id, antes=antes_txt, depois=depois_txt)
        flash(f"Servidor '{nome}' atualizado com sucesso!", "success")
    except Exception as e:
        flash(f"Erro ao atualizar servidor: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("agentes"))

@app.route("/agentes/excluir/<int:agente_id>", methods=["POST"])
@login_required
@role_required("admin")
@with_db_retry()
def agentes_excluir(agente_id):
    user_name = session["user"]["username"]
    conn = get_db_connection()
    try:
        agente = conn.execute("SELECT * FROM agentes_gcm WHERE id = ?", (agente_id,)).fetchone()
        if not agente:
            flash("Servidor não encontrado.", "warning")
            conn.close()
            return redirect(url_for("agentes"))

        taloes_count = conn.execute("SELECT COUNT(*) FROM taloes WHERE agente_gcm_id = ?", (agente_id,)).fetchone()[0]
        aits_count = conn.execute("SELECT COUNT(*) FROM ait WHERE agente_gcm_id = ?", (agente_id,)).fetchone()[0]
        
        if taloes_count > 0 or aits_count > 0:
            conn.execute("UPDATE agentes_gcm SET situacao = 'INATIVO' WHERE id = ?", (agente_id,))
            conn.commit()
            log_auditoria(user_name, "admin", "INATIVAR_AGENTE", "agentes_gcm", agente_id, justificativa="Servidor possui talões/AITs vinculados. Situação alterada para INATIVO.")
            flash(f"Servidor '{agente['nome_completo']}' possui registros vinculados ({taloes_count} talões, {aits_count} AITs) e foi marcado como INATIVO para preservar o histórico.", "warning")
        else:
            conn.execute("DELETE FROM agentes_gcm WHERE id = ?", (agente_id,))
            conn.commit()
            log_auditoria(user_name, "admin", "EXCLUIR_AGENTE", "agentes_gcm", agente_id, antes=f"{agente['nome_completo']} ({agente['matricula']})")
            flash(f"Servidor '{agente['nome_completo']}' excluído com sucesso!", "success")
    except Exception as e:
        flash(f"Erro ao excluir servidor: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("agentes"))

# --- Talões ---
@app.route("/taloes")
@login_required
def taloes():
    conn = get_db_connection()
    agentes_list = conn.execute("SELECT id, nome_completo, matricula, categoria FROM agentes_gcm WHERE situacao = 'ATIVO' ORDER BY nome_completo ASC").fetchall()
    
    taloes_query = """
    SELECT t.*, a.nome_completo as agente_nome, a.matricula as agente_matricula,
           (SELECT COUNT(*) FROM ait WHERE talao_id = t.id AND data_ait IS NOT NULL AND data_ait != '') as qtd_entregue
    FROM taloes t
    JOIN agentes_gcm a ON t.agente_gcm_id = a.id
    ORDER BY t.id DESC
    """
    taloes_rows = conn.execute(taloes_query).fetchall()
    taloes_list = []
    for r in taloes_rows:
        dict_r = dict(r)
        dict_r['qtd_faltante'] = dict_r['quantidade_calculada'] - dict_r['qtd_entregue']
        taloes_list.append(dict_r)
    conn.close()

    return render_template("taloes.html", agentes=agentes_list, taloes=taloes_list)

@app.route("/taloes/criar", methods=["POST"])
@login_required
@role_required("transporte", "dct", "admin")
@with_db_retry()
def taloes_criar():
    agente_id = request.form.get("agente_gcm_id")
    data_entrega = request.form.get("data_entrega")
    num_inicial = int(request.form.get("numero_inicial"))
    num_final = int(request.form.get("numero_final"))
    num_recibo = request.form.get("numero_recibo", "").strip()
    user_name = session["user"]["username"]

    qtd = (num_final - num_inicial) + 1
    if qtd <= 0:
        flash("Erro: O número final deve ser maior ou igual ao número inicial.", "danger")
        return redirect(url_for("taloes"))

    conn = get_db_connection()
    overlap = conn.execute("SELECT id FROM ait WHERE CAST(numero_ait AS INTEGER) BETWEEN ? AND ?", (num_inicial, num_final)).fetchone()
    if overlap:
        flash(f"Erro: Já existem AITs cadastrados na faixa {num_inicial} a {num_final}.", "danger")
        conn.close()
        return redirect(url_for("taloes"))

    cursor = conn.cursor()
    cursor.execute("""
    INSERT INTO taloes (numero_recibo, data_entrega, agente_gcm_id, numero_inicial, numero_final, quantidade_calculada, criado_por)
    VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (num_recibo, data_entrega, agente_id, num_inicial, num_final, qtd, user_name))
    talao_id = cursor.lastrowid

    for n in range(num_inicial, num_final + 1):
        num_str = str(n)
        cursor.execute("""
        INSERT INTO ait (numero_ait, talao_id, agente_gcm_id, agente_original_id, status_conferencia, criado_por, status)
        VALUES (?, ?, ?, ?, 'PENDENTE', ?, 'DCT PROCESSAR')
        """, (num_str, talao_id, agente_id, agente_id, user_name))

    conn.commit()
    conn.close()

    log_auditoria(user_name, session["user"]["setor"], "CRIAR_TALAO", "taloes", talao_id, depois=f"Faixa {num_inicial}-{num_final} ({qtd} AITs)")
    flash(f"Talão #{talao_id} cadastrado com sucesso! {qtd} AITs gerados para o responsável.", "success")
    return redirect(url_for("taloes"))

@app.route("/taloes/transferir", methods=["POST"])
@login_required
@role_required("dct", "admin")
@with_db_retry()
def taloes_transferir():
    origem_id = request.form.get("agente_origem_id")
    destino_id = request.form.get("agente_destino_id")
    t_inicial = int(request.form.get("transf_inicial"))
    t_final = int(request.form.get("transf_final"))
    motivo = request.form.get("motivo", "").strip()
    user_name = session["user"]["username"]

    if origem_id == destino_id:
        flash("Origem e Destino da transferência devem ser diferentes.", "danger")
        return redirect(url_for("taloes"))

    conn = get_db_connection()
    conn.execute("""
    UPDATE ait 
    SET agente_gcm_id = ?
    WHERE agente_gcm_id = ? AND (data_ait IS NULL OR data_ait = '')
    AND CAST(numero_ait AS INTEGER) BETWEEN ? AND ?
    """, (destino_id, origem_id, t_inicial, t_final))
    conn.commit()
    conn.close()

    log_auditoria(user_name, session["user"]["setor"], "TRANSFERIR_AITS", "ait", justificativa=motivo, depois=f"Faixa {t_inicial}-{t_final} do Agente {origem_id} -> {destino_id}")
    flash(f"AITs pendentes na faixa {t_inicial} à {t_final} transferidos com sucesso!", "success")
    return redirect(url_for("taloes"))

@app.route("/taloes/editar/<int:talao_id>", methods=["POST"])
@login_required
@role_required("dct", "admin")
@with_db_retry()
def taloes_editar(talao_id):
    num_recibo = request.form.get("numero_recibo", "").strip()
    data_entrega = request.form.get("data_entrega", "").strip()
    agente_id = request.form.get("agente_gcm_id")
    situacao = request.form.get("situacao", "ENTREGUE").strip()
    user_name = session["user"]["username"]

    conn = get_db_connection()
    try:
        old_talao = conn.execute("SELECT t.*, a.nome_completo as agente_nome FROM taloes t LEFT JOIN agentes_gcm a ON t.agente_gcm_id = a.id WHERE t.id = ?", (talao_id,)).fetchone()
        antes_txt = f"Recibo: {old_talao['numero_recibo'] or 'S/N'} | Data: {old_talao['data_entrega']} | Resp: {old_talao['agente_nome']} | Situação: {old_talao['situacao']}" if old_talao else None
        
        new_agente = conn.execute("SELECT nome_completo FROM agentes_gcm WHERE id = ?", (agente_id,)).fetchone()
        new_agente_nome = new_agente['nome_completo'] if new_agente else f"ID {agente_id}"
        depois_txt = f"Recibo: {num_recibo or 'S/N'} | Data: {data_entrega} | Resp: {new_agente_nome} | Situação: {situacao}"

        conn.execute("""
            UPDATE taloes 
            SET numero_recibo = ?, data_entrega = ?, agente_gcm_id = ?, situacao = ?
            WHERE id = ?
        """, (num_recibo, data_entrega, agente_id, situacao, talao_id))
        # Atualiza o responsável nos AITs ainda não preenchidos deste talão
        conn.execute("""
            UPDATE ait 
            SET agente_gcm_id = ?
            WHERE talao_id = ? AND (data_ait IS NULL OR data_ait = '')
        """, (agente_id, talao_id))
        conn.commit()
        log_auditoria(user_name, session["user"]["setor"], "EDITAR_TALAO", "taloes", talao_id, antes=antes_txt, depois=depois_txt)
        flash(f"Talão #{talao_id} atualizado com sucesso!", "success")
    except Exception as e:
        flash(f"Erro ao atualizar talão: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("taloes"))

@app.route("/taloes/excluir/<int:talao_id>", methods=["POST"])
@login_required
@role_required("admin")
@with_db_retry()
def taloes_excluir(talao_id):
    user_name = session["user"]["username"]
    motivo = request.form.get("motivo", "").strip() or "Exclusão administrativa de talão"
    conn = get_db_connection()
    try:
        talao = conn.execute("SELECT * FROM taloes WHERE id = ?", (talao_id,)).fetchone()
        if not talao:
            flash("Talão não encontrado.", "warning")
            conn.close()
            return redirect(url_for("taloes"))

        # Excluir AITs do talão
        del_aits = conn.execute("DELETE FROM ait WHERE talao_id = ?", (talao_id,)).rowcount
        conn.execute("DELETE FROM taloes WHERE id = ?", (talao_id,))
        conn.commit()
        log_auditoria(user_name, "admin", "EXCLUIR_TALAO", "taloes", talao_id, justificativa=motivo, depois=f"Talão #{talao_id} e {del_aits} AITs vinculados excluídos")
        flash(f"Talão #{talao_id} e seus {del_aits} AITs vinculados foram excluídos com sucesso!", "success")
    except Exception as e:
        flash(f"Erro ao excluir talão: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("taloes"))

# --- Cadastro AIT ---
@app.route("/api/check_ait_duplicada/<numero_ait>")
@login_required
def check_ait_duplicada(numero_ait):
    conn = get_db_connection()
    rec = conn.execute("SELECT id FROM ait WHERE numero_ait = ? AND data_ait IS NOT NULL AND data_ait != '' LIMIT 1", (numero_ait.strip(),)).fetchone()
    conn.close()
    if rec:
        return jsonify({"exists": True, "id": rec["id"]})
    return jsonify({"exists": False})

# --- Sincronização em Tempo Real ---
@app.route("/api/sync/version")
@login_required
def api_sync_version():
    try:
        conn = get_db_connection()
        # PRAGMA data_version retorna contador incremental modificado em commits
        data_version = conn.execute("PRAGMA data_version;").fetchone()[0]
        
        # Último log de auditoria
        last_log = conn.execute("""
            SELECT id, usuario, acao, tabela_afetada, data_hora, detalhes_depois, justificativa 
            FROM auditoria_logs 
            ORDER BY id DESC LIMIT 1
        """).fetchone()

        # Contagens rápidas para atualizar contadores e navegação
        total_aits = conn.execute("SELECT COUNT(*) FROM ait WHERE data_ait IS NOT NULL AND data_ait != ''").fetchone()[0]
        total_pendentes_dct = conn.execute("SELECT COUNT(*) FROM ait WHERE status_conferencia = 'PENDENTE' AND (data_ait IS NOT NULL AND data_ait != '')").fetchone()[0]
        total_agentes = conn.execute("SELECT COUNT(*) FROM agentes_gcm").fetchone()[0]
        total_taloes = conn.execute("SELECT COUNT(*) FROM taloes").fetchone()[0]
        
        conn.close()

        log_data = None
        if last_log:
            log_data = {
                "id": last_log["id"],
                "usuario": last_log["usuario"],
                "acao": last_log["acao"],
                "tabela": last_log["tabela_afetada"],
                "data_hora": last_log["data_hora"],
                "detalhes": last_log["detalhes_depois"] or last_log["justificativa"] or ""
            }

        return jsonify({
            "status": "ok",
            "data_version": data_version,
            "last_log": log_data,
            "counts": {
                "total_aits": total_aits,
                "pendentes_dct": total_pendentes_dct,
                "total_agentes": total_agentes,
                "total_taloes": total_taloes
            }
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/cadastro", methods=["GET", "POST"])
@app.route("/cadastro/<int:ait_id>", methods=["GET", "POST"])
@login_required
@with_db_retry()
def cadastro(ait_id=None):
    conn = get_db_connection()
    total_records = conn.execute("SELECT COUNT(*) FROM ait WHERE data_ait IS NOT NULL AND data_ait != ''").fetchone()[0]
    user_name = session["user"]["username"]
    user_setor = session["user"]["setor"]

    if request.method == "POST":
        numero_ait = request.form.get("numero_ait", "").strip()
        placa = request.form.get("placa", "").strip().upper()
        data_ait = request.form.get("data_ait", "")
        agente_mat = request.form.get("agente", "").strip()
        status = request.form.get("status", "DCT PROCESSAR")
        observacao = request.form.get("observacao", "").strip()
        data_digitacao = request.form.get("data_digitacao", "")

        if not numero_ait or not data_ait or not agente_mat:
            flash("Número AIT, Data e Código do Agente são obrigatórios!", "danger")
        else:
            agente_rec = conn.execute("SELECT id FROM agentes_gcm WHERE matricula = ?", (agente_mat,)).fetchone()
            agente_gcm_id = agente_rec["id"] if agente_rec else None

            existing = conn.execute("SELECT * FROM ait WHERE numero_ait = ?", (numero_ait,)).fetchone()

            if existing:
                if existing["status_conferencia"] == "CONFERIDO" and user_setor != "admin":
                    flash("Erro: AIT conferido pelo DCT. Edição não permitida.", "danger")
                    conn.close()
                    return redirect(url_for("cadastro", ait_id=existing["id"]))

                antes_txt = f"AIT {existing['numero_ait']} | Placa: {existing['placa'] or 'S/P'} | Data: {existing['data_ait']} | Agente: {existing['agente']} | Status: {existing['status']}"
                depois_txt = f"AIT {numero_ait} | Placa: {placa or 'S/P'} | Data: {data_ait} | Agente: {agente_mat} | Status: {status}"

                conn.execute("""
                UPDATE ait
                SET data_ait=?, agente=?, status=?, observacao=?, data_digitacao=?, placa=?, atualizado_por=?, agente_gcm_id=COALESCE(?, agente_gcm_id)
                WHERE id=?
                """, (data_ait, agente_mat, status, observacao, data_digitacao, placa, user_name, agente_gcm_id, existing["id"]))
                conn.commit()
                log_auditoria(user_name, user_setor, "EDITAR_AIT", "ait", existing["id"], antes=antes_txt, depois=depois_txt)
                flash(f"AIT #{existing['id']} salvo com sucesso!", "success")
                conn.close()
                return redirect(url_for("cadastro", ait_id=existing["id"]))
            else:
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                cursor = conn.cursor()
                cursor.execute("""
                INSERT INTO ait (numero_ait, data_ait, agente, status, observacao, data_digitacao, placa, criado_por, criado_em, agente_gcm_id, status_conferencia)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'PENDENTE')
                """, (numero_ait, data_ait, agente_mat, status, observacao, data_digitacao, placa, user_name, now_str, agente_gcm_id))
                new_id = cursor.lastrowid
                conn.commit()
                log_auditoria(user_name, user_setor, "CRIAR_AIT", "ait", new_id, depois=f"AIT {numero_ait} | Placa: {placa or 'S/P'} | Data: {data_ait} | Agente: {agente_mat}")
                flash(f"✓ AIT {numero_ait} recebido com sucesso!", "success")
                conn.close()
                return redirect(url_for("cadastro", last_agente=agente_mat, last_data=data_ait, last_dig=data_digitacao))

    record = None
    prev_id = None
    next_id = None
    is_locked = False

    if ait_id:
        record = conn.execute("SELECT * FROM ait WHERE id = ?", (ait_id,)).fetchone()
        if not record:
            flash(f"AIT #{ait_id} não encontrado.", "warning")
            conn.close()
            return redirect(url_for("cadastro"))

        if record["status_conferencia"] == "CONFERIDO" and user_setor != "admin":
            is_locked = True

        prev_row = conn.execute("SELECT MAX(id) FROM ait WHERE id < ? AND data_ait IS NOT NULL", (ait_id,)).fetchone()
        prev_id = prev_row[0] if prev_row else None
        next_row = conn.execute("SELECT MIN(id) FROM ait WHERE id > ? AND data_ait IS NOT NULL", (ait_id,)).fetchone()
        next_id = next_row[0] if next_row else None
    else:
        last_agente = request.args.get("last_agente", "")
        last_data = request.args.get("last_data", "")
        last_dig = request.args.get("last_dig", datetime.today().strftime("%Y-%m-%d"))

        record = {
            "id": "Novo", "numero_ait": "", "placa": "",
            "data_ait": last_data, "agente": last_agente,
            "status": "DCT PROCESSAR", "observacao": "", "data_digitacao": last_dig
        }

    agentes_list = conn.execute("SELECT id, nome_completo, matricula, categoria, unidade_setor FROM agentes_gcm ORDER BY categoria ASC, CAST(matricula AS INTEGER) ASC, matricula ASC").fetchall()
    conn.close()
    return render_template("cadastro.html", record=record, prev_id=prev_id, next_id=next_id, total_records=total_records, status_options=STATUS_OPTIONS, is_locked=is_locked, agentes=agentes_list)

# --- Faltantes ---
@app.route("/faltantes")
@login_required
def faltantes():
    agente_gcm_id = request.args.get("agente_gcm_id", "")
    categoria = request.args.get("categoria", "")

    where_clauses = ["(a.data_ait IS NULL OR a.data_ait = '')"]
    params = []

    if agente_gcm_id:
        where_clauses.append("a.agente_gcm_id = ?")
        params.append(agente_gcm_id)
    if categoria:
        where_clauses.append("ag.categoria = ?")
        params.append(categoria)

    where_str = " AND ".join(where_clauses)

    sql = f"""
    SELECT a.id, a.numero_ait, a.talao_id, ag.nome_completo as agente_nome, ag.matricula as agente_matricula, ag.categoria,
           t.numero_inicial as talao_num_inicial, t.numero_final as talao_num_final, t.data_entrega as talao_data_entrega
    FROM ait a
    JOIN agentes_gcm ag ON a.agente_gcm_id = ag.id
    LEFT JOIN taloes t ON a.talao_id = t.id
    WHERE {where_str}
    ORDER BY CAST(a.numero_ait AS INTEGER) ASC LIMIT 1000
    """

    conn = get_db_connection()
    faltantes_rows = conn.execute(sql, params).fetchall()
    agentes_list = conn.execute("SELECT id, nome_completo, matricula FROM agentes_gcm ORDER BY nome_completo ASC").fetchall()
    conn.close()

    faltantes_list = []
    today = date.today()
    for r in faltantes_rows:
        dict_r = dict(r)
        dias = 0
        if dict_r['talao_data_entrega']:
            try:
                dt_e = datetime.strptime(dict_r['talao_data_entrega'], "%Y-%m-%d").date()
                dias = (today - dt_e).days
            except Exception:
                dias = 0
        dict_r['dias_decorridos'] = dias
        faltantes_list.append(dict_r)

    total_faltantes = len(faltantes_list)
    agentes_com_pendencia = len(set([f['agente_matricula'] for f in faltantes_list]))

    return render_template("faltantes.html", faltantes_list=faltantes_list, agentes=agentes_list, total_faltantes=total_faltantes, agentes_com_pendencia=agentes_com_pendencia)

# --- Remessas ---
@app.route("/remessas")
@login_required
def remessas():
    conn = get_db_connection()
    remessas_list = conn.execute("SELECT * FROM remessas ORDER BY id DESC").fetchall()
    aits_disponiveis = conn.execute("SELECT COUNT(*) FROM ait WHERE data_ait IS NOT NULL AND (remessa_id IS NULL OR remessa_id = '')").fetchone()[0]
    conn.close()
    return render_template("remessas.html", remessas_list=remessas_list, aits_disponiveis=aits_disponiveis)

@app.route("/remessas/criar", methods=["POST"])
@login_required
@role_required("transporte", "dct", "admin")
@with_db_retry()
def remessas_criar():
    user_name = session["user"]["username"]
    conn = get_db_connection()

    aits = conn.execute("SELECT id FROM ait WHERE data_ait IS NOT NULL AND (remessa_id IS NULL OR remessa_id = '')").fetchall()
    if not aits:
        flash("Nenhum AIT disponível para inclusão na remessa.", "warning")
        conn.close()
        return redirect(url_for("remessas"))

    max_id = conn.execute("SELECT MAX(id) FROM remessas").fetchone()[0] or 0
    remessa_num = f"REM-{datetime.today().year}/{max_id+1:03d}"

    cursor = conn.cursor()
    cursor.execute("""
    INSERT INTO remessas (numero_remessa, criado_por, quantidade_aits, situacao)
    VALUES (?, ?, ?, 'EM_PREPARACAO')
    """, (remessa_num, user_name, len(aits)))
    remessa_id = cursor.lastrowid

    cursor.execute("""
    UPDATE ait 
    SET remessa_id = ? 
    WHERE data_ait IS NOT NULL AND data_ait != '' AND (remessa_id IS NULL OR remessa_id = '')
    """, (remessa_id,))

    conn.commit()
    conn.close()

    log_auditoria(user_name, session["user"]["setor"], "CRIAR_REMESSA", "remessas", remessa_id, depois=f"Remessa {remessa_num} com {len(aits)} AITs")
    flash(f"Remessa '{remessa_num}' criada com {len(aits)} AITs!", "success")
    return redirect(url_for("remessas"))

@app.route("/remessas/<int:remessa_id>/fechar", methods=["POST"])
@login_required
@role_required("transporte", "dct", "admin")
@with_db_retry()
def remessas_fechar(remessa_id):
    user_name = session["user"]["username"]
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_db_connection()
    conn.execute("""
    UPDATE remessas 
    SET situacao = 'FECHADA', data_fechamento = ?, data_envio_fisico = ?
    WHERE id = ?
    """, (now_str, datetime.today().strftime("%Y-%m-%d"), remessa_id))
    conn.commit()
    conn.close()

    log_auditoria(user_name, session["user"]["setor"], "FECHAR_REMESSA", "remessas", remessa_id)
    flash(f"Remessa #{remessa_id} fechada e marcada como ENVIADA!", "success")
    return redirect(url_for("remessas"))

# --- Portal Empresa de Processamento ---
@app.route("/empresa/conferencia")
@login_required
@role_required("empresa", "admin")
def empresa_conferencia():
    remessa_id = request.args.get("remessa_id")
    conn = get_db_connection()
    remessas_disponiveis = conn.execute("SELECT * FROM remessas WHERE situacao IN ('FECHADA', 'EM_CONFERENCIA', 'COM_DIVERGENCIA') ORDER BY id DESC").fetchall()

    remessa_selecionada = None
    aits_remessa = []
    aits_conferidos_count = 0

    if remessa_id:
        remessa_selecionada = conn.execute("SELECT * FROM remessas WHERE id = ?", (remessa_id,)).fetchone()
        if remessa_selecionada:
            aits_query = """
            SELECT a.*, ag.nome_completo as agente_nome, ag.matricula as agente_matricula
            FROM ait a
            LEFT JOIN agentes_gcm ag ON a.agente_gcm_id = ag.id
            WHERE a.remessa_id = ?
            ORDER BY a.id ASC
            """
            aits_remessa = conn.execute(aits_query, (remessa_id,)).fetchall()
            aits_conferidos_count = sum(1 for item in aits_remessa if item["status_conferencia"] == "CONFERIDO")

    conn.close()
    return render_template("empresa_conferencia.html", remessas_disponiveis=remessas_disponiveis, remessa_selecionada=remessa_selecionada, aits_remessa=aits_remessa, aits_conferidos_count=aits_conferidos_count)

@app.route("/empresa/conferir_item/<int:ait_id>", methods=["POST"])
@login_required
@role_required("empresa", "admin")
@with_db_retry()
def empresa_conferir_item(ait_id):
    remessa_id = request.form.get("remessa_id")
    user_name = session["user"]["username"]
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_db_connection()
    conn.execute("""
    UPDATE ait 
    SET status_conferencia = 'CONFERIDO', conferido_por = ?, conferido_em = ?, status = 'PROCESSADO DCT'
    WHERE id = ?
    """, (user_name, now_str, ait_id))

    conn.execute("UPDATE remessas SET situacao = 'EM_CONFERENCIA' WHERE id = ?", (remessa_id,))
    conn.commit()
    conn.close()

    log_auditoria(user_name, "empresa", "CONFERIR_ITEM_EMPRESA", "ait", ait_id)
    flash(f"AIT #{ait_id} confirmado com sucesso!", "success")
    return redirect(url_for("empresa_conferencia", remessa_id=remessa_id))

@app.route("/empresa/registrar_divergencia", methods=["POST"])
@login_required
@role_required("empresa", "admin")
@with_db_retry()
def empresa_registrar_divergencia():
    remessa_id = request.form.get("remessa_id")
    ait_id = request.form.get("ait_id")
    situacao_informada = request.form.get("situacao_informada")
    obs = request.form.get("observacao_empresa", "").strip()
    user_name = session["user"]["username"]

    conn = get_db_connection()
    conn.execute("""
    INSERT INTO remessa_divergencias (remessa_id, ait_id, situacao_informada, observacao_empresa, registrado_por_empresa)
    VALUES (?, ?, ?, ?, ?)
    """, (remessa_id, ait_id, situacao_informada, obs, user_name))

    conn.execute("UPDATE ait SET status_conferencia = 'DIVERGENTE' WHERE id = ?", (ait_id,))
    conn.execute("UPDATE remessas SET situacao = 'COM_DIVERGENCIA' WHERE id = ?", (remessa_id,))

    conn.commit()
    conn.close()

    log_auditoria(user_name, "empresa", "REGISTRAR_DIVERGENCIA", "remessa_divergencias", ait_id, depois=situacao_informada, justificativa=obs)
    flash("Divergência registrada com sucesso!", "warning")
    return redirect(url_for("empresa_conferencia", remessa_id=remessa_id))

# --- Divergências ---
@app.route("/divergencias")
@login_required
@role_required("dct", "admin")
def divergencias():
    conn = get_db_connection()
    sql = """
    SELECT d.*, r.numero_remessa, a.numero_ait
    FROM remessa_divergencias d
    JOIN remessas r ON d.remessa_id = r.id
    JOIN ait a ON d.ait_id = a.id
    ORDER BY d.id DESC
    """
    divs = conn.execute(sql).fetchall()
    conn.close()
    return render_template("divergencias.html", divergencias_list=divs)

@app.route("/divergencias/resolver/<int:div_id>", methods=["POST"])
@login_required
@role_required("dct", "admin")
@with_db_retry()
def divergencias_resolver(div_id):
    providencia = request.form.get("providencia_setor", "").strip()
    user_name = session["user"]["username"]
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_db_connection()
    conn.execute("""
    UPDATE remessa_divergencias 
    SET situacao_analise = 'RESOLVIDA', providencia_setor = ?, resolvido_por_setor = ?, data_hora_resolucao = ?
    WHERE id = ?
    """, (providencia, user_name, now_str, div_id))
    conn.commit()
    conn.close()

    log_auditoria(user_name, session["user"]["setor"], "RESOLVER_DIVERGENCIA", "remessa_divergencias", div_id, justificativa=providencia)
    flash("Divergência marcada como RESOLVIDA!", "success")
    return redirect(url_for("divergencias"))

# --- Módulo DCT Conferência ---
@app.route("/dct/conferencia")
@login_required
@role_required("dct", "admin")
def dct_conferencia():
    data_inicial = request.args.get("data_inicial", "")
    data_final = request.args.get("data_final", "")
    data_dig_inicial = request.args.get("data_dig_inicial", "")
    data_dig_final = request.args.get("data_dig_final", "")
    agente = request.args.get("agente", "").strip()
    status_conferencia = request.args.get("status_conferencia", "PENDENTE")

    where_clauses = ["1=1"]
    params = []

    if data_inicial and data_final:
        where_clauses.append("data_ait BETWEEN ? AND ?")
        params.extend([data_inicial, data_final])
    if data_dig_inicial and data_dig_final:
        where_clauses.append("data_digitacao BETWEEN ? AND ?")
        params.extend([data_dig_inicial, data_dig_final])
    if agente:
        where_clauses.append("agente = ?")
        params.append(agente)
    if status_conferencia != "TODOS":
        where_clauses.append("status_conferencia = ?")
        params.append(status_conferencia)

    where_str = " AND ".join(where_clauses)
    sql = f"SELECT * FROM ait WHERE {where_str} ORDER BY COALESCE(data_digitacao, criado_em, data_ait) DESC, id DESC LIMIT 500"

    conn = get_db_connection()
    records = conn.execute(sql, params).fetchall()
    conn.close()

    return render_template("dct_conferencia.html", records=records)

@app.route("/dct/aprovar_lote", methods=["POST"])
@login_required
@role_required("dct", "admin")
@with_db_retry()
def dct_aprovar_lote():
    ait_ids = request.form.getlist("ait_ids")
    novo_status = request.form.get("novo_status", "PROCESSADO DCT")
    acao = request.form.get("acao", "aprovar")
    user_name = session["user"]["username"]
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not ait_ids:
        flash("Nenhum AIT foi selecionado para ação em lote.", "warning")
        return redirect(url_for("dct_conferencia"))

    conn = get_db_connection()

    if acao == "imprimir_termo":
        records = []
        CHUNK_SIZE = 500
        for i in range(0, len(ait_ids), CHUNK_SIZE):
            chunk = ait_ids[i:i + CHUNK_SIZE]
            placeholders = ",".join(["?"] * len(chunk))
            sql = f"SELECT * FROM ait WHERE id IN ({placeholders})"
            records.extend(conn.execute(sql, chunk).fetchall())
        conn.close()
        return render_template("termo_conferencia_print.html", records=records, novo_status=novo_status)

    CHUNK_SIZE = 500
    for i in range(0, len(ait_ids), CHUNK_SIZE):
        chunk = ait_ids[i:i + CHUNK_SIZE]
        placeholders = ",".join(["?"] * len(chunk))
        sql = f"""
            UPDATE ait 
            SET status = ?, status_conferencia = 'CONFERIDO', conferido_por = ?, conferido_em = ?
            WHERE id IN ({placeholders})
        """
        conn.execute(sql, [novo_status, user_name, now_str] + chunk)
    conn.commit()
    conn.close()

    log_auditoria(user_name, session["user"]["setor"], "APROVAR_LOTE_DCT", "ait", depois=f"{len(ait_ids)} AITs -> {novo_status}")
    flash(f"Lote de {len(ait_ids)} AITs conferido e atualizado para '{novo_status}' com sucesso!", "success")
    return redirect(url_for("dct_conferencia"))

# --- Auditoria ---
@app.route("/auditoria")
@login_required
@role_required("dct", "admin")
def auditoria():
    usuario_filtro = request.args.get("usuario", "").strip()
    acao_filtro = request.args.get("acao", "").strip()
    data_inicial = request.args.get("data_inicial", "").strip()
    data_final = request.args.get("data_final", "").strip()
    busca = request.args.get("busca", "").strip()

    where_clauses = ["1=1"]
    params = []

    if usuario_filtro:
        where_clauses.append("usuario = ?")
        params.append(usuario_filtro)
    if acao_filtro:
        where_clauses.append("acao = ?")
        params.append(acao_filtro)
    if data_inicial and data_final:
        where_clauses.append("DATE(data_hora) BETWEEN ? AND ?")
        params.extend([data_inicial, data_final])
    elif data_inicial:
        where_clauses.append("DATE(data_hora) >= ?")
        params.append(data_inicial)
    elif data_final:
        where_clauses.append("DATE(data_hora) <= ?")
        params.append(data_final)

    if busca:
        where_clauses.append("(usuario LIKE ? OR acao LIKE ? OR tabela_afetada LIKE ? OR justificativa LIKE ? OR detalhes_antes LIKE ? OR detalhes_depois LIKE ? OR ip_origem LIKE ? OR hostname LIKE ?)")
        b_wild = f"%{busca}%"
        params.extend([b_wild] * 8)

    where_sql = " AND ".join(where_clauses)
    sql = f"SELECT * FROM auditoria_logs WHERE {where_sql} ORDER BY id DESC LIMIT 500"

    conn = get_db_connection()
    logs = conn.execute(sql, params).fetchall()
    usuarios_disponiveis = [r[0] for r in conn.execute("SELECT DISTINCT usuario FROM auditoria_logs WHERE usuario IS NOT NULL AND usuario != '' ORDER BY usuario ASC").fetchall()]
    acoes_disponiveis = [r[0] for r in conn.execute("SELECT DISTINCT acao FROM auditoria_logs WHERE acao IS NOT NULL AND acao != '' ORDER BY acao ASC").fetchall()]
    conn.close()

    return render_template("auditoria.html", logs=logs, usuarios_disponiveis=usuarios_disponiveis, acoes_disponiveis=acoes_disponiveis)

# --- Usuários (Admin) ---
@app.route("/usuarios")
@login_required
@role_required("admin")
def usuarios():
    conn = get_db_connection()
    usuarios_list = conn.execute("SELECT * FROM usuarios ORDER BY id ASC").fetchall()
    conn.close()
    return render_template("usuarios.html", usuarios=usuarios_list)

@app.route("/usuarios/criar", methods=["POST"])
@login_required
@role_required("admin")
@with_db_retry()
def usuarios_criar():
    username = request.form.get("username", "").strip()
    nome_completo = request.form.get("nome_completo", "").strip()
    senha = request.form.get("senha", "").strip()
    setor = request.form.get("setor", "").strip()
    vinculo = request.form.get("vinculo", "setor_publico").strip()

    if not username or not nome_completo or not senha or not setor:
        flash("Todos os campos são obrigatórios.", "danger")
        return redirect(url_for("usuarios"))

    senha_hash = generate_password_hash(senha)
    conn = get_db_connection()
    try:
        conn.execute("""
        INSERT INTO usuarios (username, nome_completo, senha_hash, setor, vinculo)
        VALUES (?, ?, ?, ?, ?)
        """, (username, nome_completo, senha_hash, setor, vinculo))
        conn.commit()
        log_auditoria(session["user"]["username"], "admin", "CRIAR_USUARIO", "usuarios", depois=f"{username} ({setor}) - {nome_completo}")
        flash(f"Usuário '{username}' cadastrado com sucesso!", "success")
    except sqlite3.IntegrityError:
        flash(f"Erro: O usuário '{username}' já existe.", "danger")
    finally:
        conn.close()

    return redirect(url_for("usuarios"))

@app.route("/usuarios/editar/<int:user_id>", methods=["POST"])
@login_required
@role_required("admin")
@with_db_retry()
def usuarios_editar(user_id):
    nome_completo = request.form.get("nome_completo", "").strip()
    setor = request.form.get("setor", "").strip()
    vinculo = request.form.get("vinculo", "setor_publico").strip()
    ativo = int(request.form.get("ativo", 1))
    nova_senha = request.form.get("nova_senha", "").strip()
    user_name = session["user"]["username"]

    if not nome_completo or not setor:
        flash("Nome completo e Setor são obrigatórios.", "danger")
        return redirect(url_for("usuarios"))

    conn = get_db_connection()
    try:
        old_user = conn.execute("SELECT * FROM usuarios WHERE id = ?", (user_id,)).fetchone()
        antes_txt = f"{old_user['nome_completo']} ({old_user['setor']}) - Ativo: {old_user['ativo']}" if old_user else None
        depois_txt = f"{nome_completo} ({setor}) - Ativo: {ativo}" + (" [Senha alterada]" if nova_senha else "")

        if nova_senha:
            senha_hash = generate_password_hash(nova_senha)
            conn.execute("""
                UPDATE usuarios 
                SET nome_completo = ?, setor = ?, vinculo = ?, ativo = ?, senha_hash = ?
                WHERE id = ?
            """, (nome_completo, setor, vinculo, ativo, senha_hash, user_id))
        else:
            conn.execute("""
                UPDATE usuarios 
                SET nome_completo = ?, setor = ?, vinculo = ?, ativo = ?
                WHERE id = ?
            """, (nome_completo, setor, vinculo, ativo, user_id))
        conn.commit()
        log_auditoria(user_name, "admin", "EDITAR_USUARIO", "usuarios", user_id, antes=antes_txt, depois=depois_txt)
        flash("Usuário atualizado com sucesso!", "success")
    except Exception as e:
        flash(f"Erro ao atualizar usuário: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("usuarios"))

@app.route("/usuarios/excluir/<int:user_id>", methods=["POST"])
@login_required
@role_required("admin")
@with_db_retry()
def usuarios_excluir(user_id):
    if user_id == session["user"]["id"]:
        flash("Você não pode excluir sua própria conta de administrador conectada.", "danger")
        return redirect(url_for("usuarios"))

    user_name = session["user"]["username"]
    conn = get_db_connection()
    try:
        target = conn.execute("SELECT username, nome_completo FROM usuarios WHERE id = ?", (user_id,)).fetchone()
        if target:
            conn.execute("DELETE FROM usuarios WHERE id = ?", (user_id,))
            conn.commit()
            log_auditoria(user_name, "admin", "EXCLUIR_USUARIO", "usuarios", user_id, antes=f"{target['nome_completo']} ({target['username']})")
            flash(f"Usuário '{target['nome_completo']}' excluído com sucesso!", "success")
        else:
            flash("Usuário não encontrado.", "warning")
    except Exception as e:
        flash(f"Erro ao excluir usuário: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("usuarios"))

# --- Exclusão de AIT (Admin) ---
@app.route("/ait/excluir/<int:ait_id>", methods=["POST"])
@login_required
@role_required("admin")
@with_db_retry()
def ait_excluir(ait_id):
    user_name = session["user"]["username"]
    motivo = request.form.get("motivo", "").strip() or "Exclusão administrativa"
    redirect_to = request.form.get("redirect_to", "consultas")

    conn = get_db_connection()
    try:
        rec = conn.execute("SELECT * FROM ait WHERE id = ?", (ait_id,)).fetchone()
        if not rec:
            flash("AIT não encontrado.", "warning")
            conn.close()
            return redirect(url_for(redirect_to))

        conn.execute("DELETE FROM ait WHERE id = ?", (ait_id,))
        conn.commit()
        log_auditoria(user_name, "admin", "EXCLUIR_AIT", "ait", ait_id, justificativa=motivo, antes=f"Nº AIT: {rec['numero_ait']}, Placa: {rec['placa']}, Status: {rec['status']}")
        flash(f"AIT #{rec['numero_ait']} excluído com sucesso!", "success")
    except Exception as e:
        flash(f"Erro ao excluir AIT: {e}", "danger")
    finally:
        conn.close()

    if redirect_to == "cadastro":
        return redirect(url_for("cadastro"))
    return redirect(request.referrer or url_for("consultas"))

# --- Backups (Admin) ---
from scripts.backup_manager import criar_backup, restaurar_backup, listar_backups, get_base_db_path, get_backup_dir
from flask import send_from_directory

@app.route("/backups")
@login_required
@role_required("admin")
def backups():
    backups_list = listar_backups()
    active_db = get_base_db_path()
    return render_template("backups.html", backups=backups_list, active_db_path=active_db)

@app.route("/backups/criar", methods=["POST"])
@login_required
@role_required("admin")
def backups_criar():
    ok, filename_or_err, path, size_mb = criar_backup(tipo="MANUAL")
    if ok:
        log_auditoria(session["user"]["username"], "admin", "CRIAR_BACKUP", "sistema", depois=f"{filename_or_err} ({size_mb} MB)")
        flash(f"Backup manual '{filename_or_err}' ({size_mb} MB) criado com sucesso!", "success")
    else:
        flash(f"Erro ao criar backup: {filename_or_err}", "danger")
    return redirect(url_for("backups"))

@app.route("/backups/restaurar/<filename>", methods=["POST"])
@login_required
@role_required("admin")
@with_db_retry()
def backups_restaurar(filename):
    confirmacao = request.form.get("confirmacao", "").strip().upper()
    if confirmacao != "RESTAURAR":
        flash("Ação cancelada: você deve digitar exatamente a palavra 'RESTAURAR' para confirmar a recuperação.", "warning")
        return redirect(url_for("backups"))

    user_name = session["user"]["username"]
    ok, msg = restaurar_backup(filename)
    if ok:
        log_auditoria(user_name, "admin", "RESTAURAR_BACKUP", "sistema", justificativa=f"Restauração executada a partir de {filename}", depois=msg)
        flash(msg, "success")
    else:
        flash(msg, "danger")
    return redirect(url_for("backups"))

@app.route("/backups/download/<filename>")
@login_required
@role_required("admin")
def backups_download(filename):
    backup_dir = get_backup_dir()
    return send_from_directory(backup_dir, filename, as_attachment=True)

from scripts.clean_database import clean_database

@app.route("/backups/resetar_banco", methods=["POST"])
@login_required
@role_required("admin")
@with_db_retry()
def backups_resetar_banco():
    confirmacao = request.form.get("confirmacao", "").strip().upper()
    if confirmacao != "ZERAR":
        flash("Ação cancelada: você deve digitar exatamente a palavra 'ZERAR' para confirmar o reset.", "warning")
        return redirect(url_for("backups"))

    user_name = session["user"]["username"]
    try:
        # Gera cópia de segurança antes do reset
        criar_backup(tipo="PRE_RESET")
        
        # Executa limpeza de todas as tabelas operacionais e servidores
        clean_database()
        
        log_auditoria(user_name, "admin", "RESET_BANCO_DADOS", "sistema", justificativa="Reset total do banco de dados executado pelo Administrador.")
        flash("Banco de dados completamente zerado com sucesso! Todos os servidores, talões, AITs e logs foram excluídos e um backup de segurança pré-reset foi salvo.", "success")
    except Exception as e:
        flash(f"Erro ao resetar banco de dados: {e}", "danger")

    return redirect(url_for("backups"))


# --- Consultas & Relatórios ---
def get_base_filter_and_params(query_type, request_args):
    if query_type == "data":
        data_inicial = request_args.get("data_inicial", "")
        data_final = request_args.get("data_final", "")
        if data_inicial and data_final:
            return "data_ait BETWEEN ? AND ?", [data_inicial, data_final]
    elif query_type == "digitacao":
        data_dig = request_args.get("data_digitacao", "")
        if data_dig:
            return "data_digitacao = ?", [data_dig]
    elif query_type == "matricula":
        agente = request_args.get("matricula", "").strip()
        if agente:
            return "(agente = ? OR agente_gcm_id IN (SELECT id FROM agentes_gcm WHERE matricula = ?))", [agente, agente]
    elif query_type == "geral":
        return "1=1", []
    return None, []

def execute_filtered_queries(conn, query_type, request_args, opcao_filtro):
    results = {}
    total_geral = 0
    where_clause, params = get_base_filter_and_params(query_type, request_args)
    if not where_clause:
        return results, total_geral

    total_geral = conn.execute(f"SELECT COUNT(*) FROM ait WHERE {where_clause}", params).fetchone()[0]

    if "qtd_mes" in opcao_filtro:
        rows = conn.execute(
            f"SELECT strftime('%Y-%m', data_ait) as mes_ano, COUNT(*) as total FROM ait WHERE {where_clause} AND data_ait IS NOT NULL AND data_ait != '' GROUP BY mes_ano ORDER BY mes_ano", 
            params
        ).fetchall()
        results["qtd_mes"] = [dict(row) for row in rows]

    if "qtd_agentes" in opcao_filtro or "dist_agentes" in opcao_filtro:
        rows = conn.execute(
            f"SELECT agente, COUNT(*) as total FROM ait WHERE {where_clause} GROUP BY agente ORDER BY total DESC", 
            params
        ).fetchall()
        results["agentes"] = [dict(row) for row in rows]

    if "tipo_mes" in opcao_filtro:
        rows = conn.execute(
            f"SELECT strftime('%Y-%m', data_ait) as mes_ano, status, COUNT(*) as total FROM ait WHERE {where_clause} AND data_ait IS NOT NULL AND data_ait != '' GROUP BY mes_ano, status ORDER BY mes_ano, total DESC", 
            params
        ).fetchall()
        results["tipo_mes"] = [dict(row) for row in rows]

    if "lista" in opcao_filtro:
        limit_clause = " LIMIT 1000" if query_type == "geral" else ""
        rows = conn.execute(f"SELECT * FROM ait WHERE {where_clause} ORDER BY id DESC{limit_clause}", params).fetchall()
        results["lista"] = [dict(row) for row in rows]

    return results, total_geral

@app.route("/consultas")
@login_required
def consultas():
    query_type = request.args.get("tipo", "")
    opcao_filtro = request.args.getlist("opcao_filtro") or ["lista"]
    results = {}
    total_geral = 0
    if query_type:
        conn = get_db_connection()
        try:
            results, total_geral = execute_filtered_queries(conn, query_type, request.args, opcao_filtro)
        except Exception as e:
            flash(f"Erro ao executar consulta: {e}", "danger")
        finally:
            conn.close()
    return render_template("consultas.html", results=results, query_type=query_type, opcao_filtro=opcao_filtro, total_geral=total_geral)

@app.route("/relatorios")
@login_required
def relatorios():
    return render_template("relatorios.html")

@app.route("/relatorios/geral")
@login_required
def relatorio_general():
    opcao_filtro = request.args.getlist("opcao_filtro") or ["lista"]
    conn = get_db_connection()
    results, total_geral = execute_filtered_queries(conn, "geral", request.args, opcao_filtro)
    conn.close()
    return render_template("relatorio_print.html", title="Relatório Geral de Autos de Infração", results=results, subtitle="Exibindo todos os registros consolidados", opcao_filtro=opcao_filtro, total_geral=total_geral)

@app.route("/relatorios/data")
@login_required
def relatorio_data():
    data_inicial = request.args.get("data_inicial", "")
    data_final = request.args.get("data_final", "")
    opcao_filtro = request.args.getlist("opcao_filtro") or ["lista"]
    if not data_inicial or not data_final:
        flash("Informe o intervalo de datas.", "danger")
        return redirect(url_for("relatorios"))
    conn = get_db_connection()
    results, total_geral = execute_filtered_queries(conn, "data", request.args, opcao_filtro)
    conn.close()
    dt_i = datetime.strptime(data_inicial, "%Y-%m-%d").strftime("%d/%m/%Y")
    dt_f = datetime.strptime(data_final, "%Y-%m-%d").strftime("%d/%m/%Y")
    return render_template("relatorio_print.html", title="Relatório por Período de Infração", results=results, subtitle=f"Período: {dt_i} até {dt_f}", opcao_filtro=opcao_filtro, total_geral=total_geral)

@app.route("/relatorios/digitacao")
@login_required
def relatorio_digitacao():
    data_dig = request.args.get("data_digitacao", "")
    opcao_filtro = request.args.getlist("opcao_filtro") or ["lista"]
    if not data_dig:
        flash("Informe a data de digitação.", "danger")
        return redirect(url_for("relatorios"))
    conn = get_db_connection()
    results, total_geral = execute_filtered_queries(conn, "digitacao", request.args, opcao_filtro)
    conn.close()
    dt_d = datetime.strptime(data_dig, "%Y-%m-%d").strftime("%d/%m/%Y")
    return render_template("relatorio_print.html", title="Relatório por Data da Digitação", results=results, subtitle=f"Data da Digitação: {dt_d}", opcao_filtro=opcao_filtro, total_geral=total_geral)

# --- Exportação de Dados (Excel / CSV) ---
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Response

@app.route("/consultas/exportar")
@login_required
def consultas_exportar():
    fmt = request.args.get("formato", "excel")
    query_type = request.args.get("tipo", "geral")
    opcao_filtro = ["lista"]
    
    conn = get_db_connection()
    results, total_geral = execute_filtered_queries(conn, query_type, request.args, opcao_filtro)
    conn.close()
    
    rows = results.get("lista", [])
    filename_prefix = f"triagem_ait_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow(["ID", "Número AIT", "Placa", "Data AIT", "Agente/Matrícula", "Status", "Status Conferência", "Remessa ID"])
        for r in rows:
            writer.writerow([r.get("id"), r.get("numero_ait"), r.get("placa"), r.get("data_ait"), r.get("agente"), r.get("status"), r.get("status_conferencia"), r.get("remessa_id")])
        
        csv_bytes = ("\ufeff" + output.getvalue()).encode('utf-8-sig')
        return Response(csv_bytes, mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={filename_prefix}.csv"})
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AITs Exportados"
        
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        headers = ["ID", "Número AIT", "Placa", "Data AIT", "Agente/Matrícula", "Status", "Status Conferência", "Remessa ID"]
        ws.append(headers)
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for r in rows:
            ws.append([r.get("id"), r.get("numero_ait"), r.get("placa"), r.get("data_ait"), r.get("agente"), r.get("status"), r.get("status_conferencia"), r.get("remessa_id")])
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        return Response(stream.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename_prefix}.xlsx"})

@app.route("/faltantes/exportar")
@login_required
def faltantes_exportar():
    fmt = request.args.get("formato", "excel")
    conn = get_db_connection()
    
    taloes = conn.execute("SELECT t.*, ag.nome_completo as agente_nome, ag.matricula as agente_matricula FROM taloes t LEFT JOIN agentes_gcm ag ON t.agente_gcm_id = ag.id ORDER BY t.id ASC").fetchall()
    aits_existentes = set([row[0] for row in conn.execute("SELECT numero_ait FROM ait WHERE numero_ait IS NOT NULL").fetchall()])
    conn.close()

    today = date.today()
    faltantes_list = []

    for t in taloes:
        inicio = t["numero_inicial"]
        fim = t["numero_final"]
        for num in range(inicio, fim + 1):
            str_num = str(num)
            if str_num not in aits_existentes:
                dict_r = dict(t)
                dict_r['numero_ait_faltante'] = str_num
                dict_r['talao_recibo'] = t['numero_recibo']
                dict_r['talao_data_entrega'] = t['data_entrega']
                dias = 0
                if dict_r['talao_data_entrega']:
                    try:
                        dt_e = datetime.strptime(dict_r['talao_data_entrega'], "%Y-%m-%d").date()
                        dias = (today - dt_e).days
                    except Exception:
                        dias = 0
                dict_r['dias_decorridos'] = dias
                faltantes_list.append(dict_r)

    filename_prefix = f"triagem_ait_faltantes_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow(["Nº AIT Faltante", "Agente Responsável", "Matrícula", "Recibo Talão", "Data Entrega", "Dias Omissão"])
        for r in faltantes_list:
            writer.writerow([r.get("numero_ait_faltante"), r.get("agente_nome"), r.get("agente_matricula"), r.get("talao_recibo"), r.get("talao_data_entrega"), r.get("dias_decorridos")])
        
        csv_bytes = ("\ufeff" + output.getvalue()).encode('utf-8-sig')
        return Response(csv_bytes, mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={filename_prefix}.csv"})
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AITs Faltantes"
        
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        headers = ["Nº AIT Faltante", "Agente Responsável", "Matrícula", "Recibo Talão", "Data Entrega", "Dias Omissão"]
        ws.append(headers)
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for r in faltantes_list:
            ws.append([r.get("numero_ait_faltante"), r.get("agente_nome"), r.get("agente_matricula"), r.get("talao_recibo"), r.get("talao_data_entrega"), r.get("dias_decorridos")])
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        return Response(stream.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename_prefix}.xlsx"})

if __name__ == "__main__":
    print("Iniciando servidor local Triagem AIT v1.1 na porta 5000...")
    app.run(host="0.0.0.0", port=5000, debug=True)
